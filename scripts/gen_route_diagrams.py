# -*- coding: utf-8 -*-
"""답십리 행로표 .xlsm → src/data/routeDiagrams.ts 자동 생성 + 검증 몽타주.
셀(열번·시각·km)=위치, 도형 cxnSp=연결선. (col,row) 균일 그리드 → 압축 SVG."""
import openpyxl, re, datetime, zipfile, math
from xml.etree import ElementTree as ET

XL=r"D:/다이아 작업/답십리사업소_행로표_260322_210054(엑셀)/답십리사업소_행로표_260322_210054(엑셀).xlsm"
NS={'xdr':'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
    'a':'http://schemas.openxmlformats.org/drawingml/2006/main'}
COLEMU=914400; ROWEMU=209550
VARIANTS=['평일','휴일','평평','평휴','휴평','휴휴']

def sheet_to_drawing(z):
    wbrels=z.read('xl/_rels/workbook.xml.rels').decode('utf8')
    wb_xml=z.read('xl/workbook.xml').decode('utf8')
    name2rid=dict(re.findall(r'<sheet name="([^"]+)"[^>]*r:id="(rId\d+)"', wb_xml))
    rid2sheet=dict(re.findall(r'Id="(rId\d+)"[^>]*Target="worksheets/(sheet\d+\.xml)"', wbrels))
    out={}
    for name,rid in name2rid.items():
        sf=rid2sheet.get(rid)
        if not sf: continue
        sn=sf.replace('.xml','')
        try:
            shrels=z.read(f'xl/worksheets/_rels/{sn}.xml.rels').decode('utf8')
            d=re.search(r'Target="\.\./drawings/(drawing\d+\.xml)"', shrels)
            out[name]=(sf, d.group(1) if d else None)
        except KeyError: out[name]=(sf,None)
    return out

def parse_lines(z, draw):
    root=ET.fromstring(z.read(f'xl/drawings/{draw}').decode('utf8')); lines=[]
    for anc in root.findall('xdr:twoCellAnchor',NS):
        if anc.find('.//xdr:cxnSp',NS) is None: continue
        def cr(tag):
            e=anc.find(tag,NS)
            return (int(e.find('xdr:col',NS).text)+int(e.find('xdr:colOff',NS).text)/COLEMU,
                    int(e.find('xdr:row',NS).text)+int(e.find('xdr:rowOff',NS).text)/ROWEMU)
        lines.append((cr('xdr:from'),cr('xdr:to')))
    return lines

def fmt(v):
    # 일부 열번/시각이 timedelta 로 오포맷됨: "5025 days"(=열번), "6:54:30"(=시각)
    if isinstance(v, datetime.timedelta):
        if 1000<=v.days<=5999: return str(v.days)          # 열번
        s=v.total_seconds(); return f"{int(s//3600)%24:02d}:{int((s%3600)//60):02d}"  # 시:분
    if isinstance(v,(datetime.time,datetime.datetime)): return v.strftime("%H:%M")
    return str(v).strip()

def is_depot(tn): 
    n=int(tn); return 1000<=n<=1599 or 2000<=n<=2999

def extract(ws, lines):
    rows=list(ws.iter_rows(min_row=1,max_row=ws.max_row,values_only=True))
    numrows=[(i+1,int(str(r[0]).strip())) for i,r in enumerate(rows)
             if r and r[0] is not None and re.fullmatch(r'\d{1,2}',str(r[0]).strip())]
    dias={}
    for idx,(nr,dnum) in enumerate(numrows):
        r0=nr-2; r1=(numrows[idx+1][0]-3) if idx+1<len(numrows) else min(nr+20,len(rows))
        pills=[]; times=[]; kmrows=[]
        for r in range(r0,r1+1):
            if r<1 or r>len(rows): continue
            for c,v in enumerate(rows[r-1],1):
                if v is None: continue
                s=fmt(v)
                if re.fullmatch(r'[1-9]\d{3}',s) and c>=4: pills.append((s,c,r))  # 9599 등 9xxx 포함
                elif re.fullmatch(r'\d{1,2}:\d{2}',s) and c>=3: times.append((s,c,r))
                elif 'km' in s and c<=2: kmrows.append((s.replace(' ',''),r))  # 실제 주행키로 값
        if not pills: continue
        blines=[(a,b) for a,b in lines if r0-0.5<=a[1]<=r1+0.5 and r0-0.5<=b[1]<=r1+0.5]
        dias[dnum]={'pills':pills,'times':times,'kms':sorted(set(kmrows),key=lambda x:x[1]),'lines':blines,'r0':r0,'r1':r1}
    return dias

def seg_split(pills):
    prs=sorted(set(r for _,_,r in pills)); segs=[]; cur=[prs[0]]
    for r in prs[1:]:
        if r-cur[-1]>=3: segs.append((cur[0],cur[-1])); cur=[r]
        else: cur.append(r)
    segs.append((cur[0],cur[-1])); return segs

def seg_dist(pt, s):
    (ax,ay),(bx,by)=s; px,py=pt; dx,dy=bx-ax,by-ay
    if dx==0 and dy==0: return ((px-ax)**2+(py-ay)**2)**0.5
    tt=max(0.0,min(1.0,((px-ax)*dx+(py-ay)*dy)/(dx*dx+dy*dy)))
    return ((px-(ax+tt*dx))**2+(py-(ay+tt*dy))**2)**0.5

def tomin(s):
    try:
        h,m=s.split(':'); return int(h)*60+int(m)
    except: return 0

def render(d):
    YS=15; TRUE=36; out=[]  # 원본 실척(≈36px/열), 압축 없음 → 넘치면 가로 스크롤(Option 1)
    for (ss,se) in seg_split(d['pills']):
        lo,hi=ss-1,se   # 마지막 알약 행까지만 (그 아래 footer 요약행 배제)
        p=[x for x in d['pills'] if ss<=x[2]<=se]
        t=[x for x in d['times'] if lo<=x[2]<=hi]
        ln=[(a,b) for a,b in d['lines'] if lo<=a[1]<=hi and lo<=b[1]<=hi]
        # ── 편승 = '가로 두 줄 평행선' 만 제거 (하남선 등 실제 운행선은 전부 유지) ──
        hz=[i for i,(a,b) in enumerate(ln) if abs(a[1]-b[1])<0.3]  # 수평선 인덱스
        pdrm=set(); pend=[]
        for ii in range(len(hz)):
            for jj in range(ii+1,len(hz)):
                a1,b1=ln[hz[ii]]; a2,b2=ln[hz[jj]]
                if 0.05<abs(a1[1]-a2[1])<0.5:            # 다른 행이지만 아주 가까운 두 수평선
                    x1=sorted([a1[0],b1[0]]); x2=sorted([a2[0],b2[0]])
                    if min(x1[1],x2[1])-max(x1[0],x2[0])>2:  # x구간 상당 겹침 → 편승 두줄
                        pdrm|={hz[ii],hz[jj]}; pend+=[a1,b1,a2,b2]
        ln=[s for i,s in enumerate(ln) if i not in pdrm]
        # 편승 선 끝점 근처 시각 제거 (예: 22:26). 실제 운행 시각은 유지
        t=[x for x in t if not any(abs(x[1]-e[0])<2 and abs(x[2]-e[1])<1.3 for e in pend)]
        # 중복 열번 알약 → 하나만(원본 1개). 선은 유지(실제 운행선 보존)
        by={}
        for tn,c,r in p:
            if tn not in by or c<by[tn][1]: by[tn]=(tn,c,r)
        p=list(by.values())
        # 끝점 0.5 스냅 (미세 간격 이음). 추가 브리지선은 그리지 않음(원본에 없는 선 금지)
        sn=lambda v: round(v*2)/2
        ln=[((sn(a[0]),sn(a[1])),(sn(b[0]),sn(b[1]))) for a,b in ln]
        ln=[(a,b) for a,b in ln if a!=b]
        # ── 동일 행 수평선 병합: 시각 span ≤90분이면 왼쪽으로 이어붙여 하나의 선(끊김·하남행 detach 제거) ──
        rt={}
        for s,c,r in t: rt.setdefault(round(r),[]).append(tomin(s))
        hzr={}
        for i,(a,b) in enumerate(ln):
            if abs(a[1]-b[1])<0.3: hzr.setdefault(round(a[1]),[]).append(i)
        drop=set(); addln=[]
        for row,idxs in hzr.items():
            tms=rt.get(row,[])
            if len(idxs)>=2 and tms and (max(tms)-min(tms))<=120:
                xs=[ln[i][0][0] for i in idxs]+[ln[i][1][0] for i in idxs]
                yv=ln[idxs[0]][0][1]
                addln.append(((min(xs),yv),(max(xs),yv))); drop|=set(idxs)
        ln=[s for i,s in enumerate(ln) if i not in drop]+addln
        # ── 원본 실척(압축 없음): x = 열 * TRUE ──
        cols=[c for _,c,_ in p]+[c for _,c,_ in t]+[a[0] for a,b in ln]+[b[0] for a,b in ln]
        mincol=min(cols)
        BASE=50
        X=lambda col: round(BASE+(col-mincol)*TRUE,1)
        minrow=min([r for _,_,r in p]+[r for _,_,r in t])
        Y=lambda r: round((r-minrow)*YS+22,1)
        # hs=회송열차(5900번대) → 영업열차와 구분 표시
        pills=[{'t':tn,'x':round(X(c)-23,1),'y':round(Y(r)-11,1),'hs':(5900<=int(tn)<=5999)} for tn,c,r in p]
        # 출발시각 = 근무에서 가장 이른 시각(=g.d) 을 크게
        dep=min(t,key=lambda x:tomin(x[0])) if t else None
        epts=[(a[0],a[1]) for a,b in ln]+[(b[0],b[1]) for a,b in ln]  # 선 끝점
        texts=[]
        for s,c,r in t:
            cc=c
            # 시각을 같은 행 선 끝점에 스냅 → 원본처럼 선 끝에 붙임(멀리 안 그림)
            cand=[e for e in epts if abs(e[1]-r)<0.6]
            if cand:
                ne=min(cand,key=lambda e:abs(e[0]-cc))
                if abs(ne[0]-cc)<3.5: cc=ne[0]
            rowp=[pp for pp in p if pp[2]==r]
            nb=min(rowp or p, key=lambda pp:abs(pp[1]-c))
            anc='end' if c<nb[1] else 'start'   # 좌/우는 원래 열 기준
            xx=X(cc)+(5 if anc=='start' else -5)
            # 알약(열번) 안쪽으로 들어가지 않게 밖으로 밀어냄 — 기지 출고/입고 시간 가독(중요)
            pl=X(nb[1])-23; pr=X(nb[1])+23
            xx=max(xx,pr+4) if anc=='start' else min(xx,pl-4)
            tx={'t':s,'x':round(xx,1),'y':round(Y(r)+4,1),'anchor':anc,'kind':'time'}
            if dep and (s,c,r)==dep: tx['dep']=True
            texts.append(tx)
        kmv=[v for v,rr in d['kms'] if lo<=rr<=hi]
        nodes=[]
        firstrow=min(r for _,_,r in p)
        f=[x for x in p if x[2]==firstrow][0]
        if is_depot(f[0]):
            # 출고 시작점 ● = 그 행 '가장 이른 시각(출발)' 위치. 2xxx=왼쪽 출발, 1xxx=오른쪽 출발 자동 반영
            ft=[x for x in t if x[2]==firstrow]
            sc=min(ft,key=lambda x:tomin(x[0]))[1] if ft else f[1]
            nodes.append({'x':X(sc),'y':round(Y(firstrow)-22,1)})
            ln=ln+[((sc,firstrow-1.4),(sc,firstrow-0.7))]
        paths=[f"M{X(a[0])},{Y(a[1])} L{X(b[0])},{Y(b[1])}" for a,b in ln]
        xs=[X(c) for _,c,_ in p]+[X(c) for _,c,_ in t]
        cxmin=min(xs)-50; cxmax=max(xs)+44
        # 선(path)·노드 y도 포함해 viewBox 잡음 → 아래로 뻗은 선 잘림 방지
        ys=[Y(r) for _,_,r in p]+[Y(r) for _,_,r in t]
        ys+=[Y(a[1]) for a,b in ln]+[Y(b[1]) for a,b in ln]+[n['y'] for n in nodes]
        top=min(ys)-14; kmy=round(Y(minrow)-14,1)
        if kmv: texts.append({'t':kmv[0],'x':round(cxmin+8,1),'y':kmy,'kind':'km'}); top=min(top,kmy-13)
        vy=round(top-6); vh=max(ys)-vy+16
        natural=cxmax-cxmin
        VW=max(260,round(natural))  # 1:1 렌더용 자연 폭 (넘치면 컴포넌트에서 가로 스크롤)
        vx=round(cxmin-(VW-natural)/2)
        out.append({'viewBox':[vx,vy,VW,round(vh)],'paths':paths,'nodes':nodes,'pills':pills,'texts':texts})
    return out

def main():
    z=zipfile.ZipFile(XL); s2d=sheet_to_drawing(z)
    wb=openpyxl.load_workbook(XL, data_only=True, read_only=True)
    data={}
    for v in VARIANTS:
        sf,draw=s2d[v]
        if not draw: print(v,"no drawing"); continue
        dias=extract(wb[v], parse_lines(z,draw))
        data[v]={dn:render(d) for dn,d in dias.items()}
        print(v, "교번", len(dias), sorted(dias)[:3],"...",sorted(dias)[-2:])
    import json
    open(r"scripts/_route_data.json","w",encoding="utf8").write(json.dumps(data,ensure_ascii=False))
    # 검증: 79 seg
    for i,s in enumerate(data['평평'][79]):
        print(f"[79 seg{i}]", [(pp['t']) for pp in s['pills']], "vb",s['viewBox'])
main()
